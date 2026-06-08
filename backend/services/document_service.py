import re
import shutil
import subprocess
import tempfile
import zipfile
from pathlib import Path
from xml.etree import ElementTree


ALLOWED_EXTENSIONS = {"pdf", "doc", "docx", "txt", "md", "markdown", "pptx", "xlsx"}
SUPPORTED_FILE_TYPES_LABEL = "PDF、Word(doc/docx)、TXT、Markdown(md)、PPTX、XLSX 文件"


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def extract_text(file_path):
    path = Path(file_path)
    extension = path.suffix.lower().lstrip(".")
    if extension == "pdf":
        return _extract_pdf(path)
    if extension == "doc":
        return _extract_doc(path)
    if extension == "docx":
        return _extract_docx(path)
    if extension == "txt":
        return _read_text_file(path)
    if extension in {"md", "markdown"}:
        return _extract_markdown(path)
    if extension == "pptx":
        return _extract_pptx(path)
    if extension == "xlsx":
        return _extract_xlsx(path)
    raise ValueError("Unsupported file type")


def split_text(text, chunk_size=900, overlap=120):
    cleaned = re.sub(r"\r\n?", "\n", text).strip()
    if not cleaned:
        return []

    paragraphs = [item.strip() for item in re.split(r"\n{2,}", cleaned) if item.strip()]
    chunks = []
    current = ""
    for paragraph in paragraphs:
        if len(current) + len(paragraph) + 2 <= chunk_size:
            current = f"{current}\n\n{paragraph}".strip()
            continue
        if current:
            chunks.extend(_slice_long_text(current, chunk_size, overlap))
        current = paragraph
    if current:
        chunks.extend(_slice_long_text(current, chunk_size, overlap))
    return chunks


def _extract_pdf(path):
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise ValueError("解析 PDF 需要安装 pypdf") from exc

    reader = PdfReader(str(path))
    pages = [page.extract_text() or "" for page in reader.pages]
    return "\n\n".join(pages)


def _extract_docx(path):
    try:
        from docx import Document
    except ImportError:
        return _extract_docx_xml(path)

    document = Document(str(path))
    paragraphs = [paragraph.text for paragraph in document.paragraphs if paragraph.text.strip()]
    text = "\n\n".join(paragraphs)
    return text or _extract_docx_xml(path)


def _extract_docx_xml(path):
    parts = []
    with zipfile.ZipFile(path) as archive:
        names = sorted(
            [
                name
                for name in archive.namelist()
                if name == "word/document.xml"
                or re.match(r"word/(header|footer|footnotes|endnotes|comments)\d*\.xml$", name)
            ],
            key=lambda name: 0 if name == "word/document.xml" else 1,
        )
        for name in names:
            texts = _extract_xml_text_nodes(archive.read(name))
            if texts:
                parts.append("\n".join(texts))
    return "\n\n".join(parts)


def _extract_doc(path):
    errors = []
    extractors = (
        _extract_doc_with_word,
        _extract_doc_with_libreoffice,
        _extract_doc_with_antiword,
    )
    for extractor in extractors:
        try:
            text = extractor(path)
            if text.strip():
                return text
        except Exception as exc:
            errors.append(f"{extractor.__name__}: {exc}")
    raise ValueError(
        "DOC 文件需要本机安装 Microsoft Word、LibreOffice/soffice 或 antiword 才能解析。"
        f"当前转换失败：{'; '.join(errors)}"
    )


def _extract_doc_with_word(path):
    try:
        import pythoncom
        import win32com.client
    except ImportError as exc:
        raise RuntimeError("未安装 pywin32") from exc

    pythoncom.CoInitialize()
    word = None
    document = None
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / f"{path.stem}.docx"
            word = win32com.client.DispatchEx("Word.Application")
            word.Visible = False
            word.DisplayAlerts = 0
            document = word.Documents.Open(str(path.resolve()), ReadOnly=True, ConfirmConversions=False)
            document.SaveAs2(str(output_path), FileFormat=16)
            document.Close(False)
            document = None
            return _extract_docx(output_path)
    finally:
        if document is not None:
            document.Close(False)
        if word is not None:
            word.Quit()
        pythoncom.CoUninitialize()


def _extract_doc_with_libreoffice(path):
    executable = shutil.which("soffice") or shutil.which("libreoffice")
    if not executable:
        raise RuntimeError("未找到 LibreOffice/soffice")

    with tempfile.TemporaryDirectory() as temp_dir:
        completed = subprocess.run(
            [
                executable,
                "--headless",
                "--convert-to",
                "docx",
                "--outdir",
                temp_dir,
                str(path.resolve()),
            ],
            capture_output=True,
            timeout=90,
            check=False,
        )
        if completed.returncode != 0:
            message = completed.stderr.decode("utf-8", errors="ignore").strip()
            raise RuntimeError(message or "LibreOffice 转换失败")
        converted_files = list(Path(temp_dir).glob("*.docx"))
        if not converted_files:
            raise RuntimeError("LibreOffice 未生成 docx 文件")
        return _extract_docx(converted_files[0])


def _extract_doc_with_antiword(path):
    executable = shutil.which("antiword")
    if not executable:
        raise RuntimeError("未找到 antiword")
    completed = subprocess.run(
        [executable, str(path.resolve())],
        capture_output=True,
        timeout=60,
        check=False,
    )
    if completed.returncode != 0:
        message = completed.stderr.decode("utf-8", errors="ignore").strip()
        raise RuntimeError(message or "antiword 解析失败")
    return completed.stdout.decode("utf-8", errors="ignore")


def _extract_markdown(path):
    text = _read_text_file(path)
    text = re.sub(r"```[\s\S]*?```", lambda item: item.group(0).replace("```", ""), text)
    text = re.sub(r"!\[([^\]]*)\]\([^)]+\)", r"\1", text)
    text = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", text)
    text = re.sub(r"^#{1,6}\s*", "", text, flags=re.MULTILINE)
    text = re.sub(r"`([^`]+)`", r"\1", text)
    text = re.sub(r"[*_~>]+", "", text)
    return text


def _extract_pptx(path):
    parts = []
    with zipfile.ZipFile(path) as archive:
        slide_names = sorted(
            [
                name
                for name in archive.namelist()
                if name.startswith("ppt/slides/slide") and name.endswith(".xml")
            ],
            key=_natural_key,
        )
        for index, name in enumerate(slide_names, start=1):
            texts = _extract_xml_text_nodes(archive.read(name))
            if texts:
                parts.append(f"幻灯片 {index}\n" + "\n".join(texts))

        note_names = sorted(
            [
                name
                for name in archive.namelist()
                if name.startswith("ppt/notesSlides/notesSlide") and name.endswith(".xml")
            ],
            key=_natural_key,
        )
        for index, name in enumerate(note_names, start=1):
            texts = _extract_xml_text_nodes(archive.read(name))
            if texts:
                parts.append(f"备注 {index}\n" + "\n".join(texts))
    return "\n\n".join(parts)


def _extract_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("解析 XLSX 需要安装 openpyxl") from exc

    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    parts = []
    try:
        for sheet in workbook.worksheets:
            rows = []
            for row in sheet.iter_rows(values_only=True):
                cells = [str(value).strip() for value in row if value is not None and str(value).strip()]
                if cells:
                    rows.append("\t".join(cells))
            if rows:
                parts.append(f"工作表：{sheet.title}\n" + "\n".join(rows))
    finally:
        workbook.close()
    return "\n\n".join(parts)


def _extract_xml_text_nodes(xml_bytes):
    root = ElementTree.fromstring(xml_bytes)
    texts = []
    for node in root.iter():
        if node.tag.endswith("}t") and node.text and node.text.strip():
            texts.append(node.text.strip())
    return texts


def _read_text_file(path):
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="ignore")


def _natural_key(value):
    return [int(item) if item.isdigit() else item for item in re.split(r"(\d+)", value)]


def _slice_long_text(text, chunk_size, overlap):
    if len(text) <= chunk_size:
        return [text]
    pieces = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        piece = text[start:end].strip()
        if piece:
            pieces.append(piece)
        start = max(end - overlap, start + 1)
    return pieces
